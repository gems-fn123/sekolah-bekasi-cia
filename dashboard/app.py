from __future__ import annotations

import io
import re
from typing import Any

import openpyxl
import pandas as pd
import plotly.express as px
import requests
import streamlit as st


st.set_page_config(page_title="Preschool Budget Dashboard", layout="wide")


st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;700&family=Source+Serif+4:wght@500;700&display=swap');

:root {
  --bg-a: #fffdf3;
  --bg-b: #f0fbf7;
  --ink: #1f2937;
  --card: #ffffff;
  --line: #d1d5db;
  --brand: #0f766e;
  --brand-soft: #e6fffa;
}

html, body, [class*="css"] {
  font-family: 'Space Grotesk', sans-serif;
  color: var(--ink);
}

.stApp {
  background:
    radial-gradient(circle at 12% 10%, #fff7d6 0%, rgba(255, 247, 214, 0) 40%),
    radial-gradient(circle at 88% 18%, #dbf4ff 0%, rgba(219, 244, 255, 0) 36%),
    linear-gradient(160deg, var(--bg-a) 0%, var(--bg-b) 100%);
}

h1, h2, h3 {
  font-family: 'Source Serif 4', serif;
}

.kpi-card {
  background: var(--card);
  border: 1px solid var(--line);
  border-radius: 16px;
  padding: 14px 16px;
  box-shadow: 0 8px 24px rgba(15, 23, 42, 0.05);
}

.kpi-label {
  font-size: 0.86rem;
  color: #4b5563;
}

.kpi-value {
  font-size: 1.36rem;
  font-weight: 700;
  margin-top: 4px;
}

.kpi-hint {
  margin-top: 2px;
  font-size: 0.8rem;
  color: #6b7280;
}

.block-label {
  font-size: 0.86rem;
  color: #4b5563;
  margin-bottom: 2px;
}

.block-value {
  font-size: 1.1rem;
  font-weight: 700;
}

.small-muted {
  color: #6b7280;
  font-size: 0.85rem;
}
</style>
""",
    unsafe_allow_html=True,
)


def idr(value: float) -> str:
    return f"Rp {value:,.0f}".replace(",", ".")


def to_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace("Rp", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def extract_drive_file_id(url: str) -> str | None:
    patterns = [
        r"/d/([a-zA-Z0-9_-]+)",
        r"id=([a-zA-Z0-9_-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None


@st.cache_data(show_spinner=False)
def download_drive_xlsx(share_url: str) -> bytes:
    file_id = extract_drive_file_id(share_url)
    if not file_id:
        raise ValueError("Google Drive link format not recognized. Use a shared file URL.")
    direct_url = f"https://drive.google.com/uc?export=download&id={file_id}"
    response = requests.get(direct_url, timeout=30)
    response.raise_for_status()
    return response.content


def load_workbook(source_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(source_bytes), data_only=True)


def get_month_sheets(workbook) -> list[str]:
    month_sheets: list[str] = []
    for sheet in workbook.worksheets:
        title = sheet.title.strip().lower()
        if title == "budget template":
            continue
        if sheet.max_row <= 1:
            continue
        if str(sheet["B2"].value).strip().lower() == "income":
            month_sheets.append(sheet.title)
    return month_sheets


def parse_budget_table(workbook) -> pd.DataFrame:
    if "Budget Template" not in workbook.sheetnames:
        return pd.DataFrame()

    sheet = workbook["Budget Template"]
    header_row = None
    for row in range(1, min(sheet.max_row, 80) + 1):
        first_cell = str(sheet.cell(row, 1).value).strip().lower()
        if first_cell == "kategori":
            header_row = row
            break

    if header_row is None:
        return pd.DataFrame()

    records = []
    for row in range(header_row + 1, sheet.max_row + 1):
        category = sheet.cell(row, 1).value
        if category is None:
            continue

        if str(category).strip().upper() == "TOTAL":
            break

        records.append(
            {
                "Category": str(category).strip(),
                "Budget": to_number(sheet.cell(row, 2).value),
                "Actual": to_number(sheet.cell(row, 3).value),
                "Remaining": to_number(sheet.cell(row, 4).value),
                "UsedPct": to_number(sheet.cell(row, 5).value),
                "Status": str(sheet.cell(row, 6).value or "").strip(),
                "Type": str(sheet.cell(row, 7).value or "").strip(),
                "Notes": str(sheet.cell(row, 8).value or "").strip(),
            }
        )

    return pd.DataFrame(records)


def section_start(sheet, row: int) -> bool:
    if row + 1 > sheet.max_row:
        return False
    header = sheet.cell(row, 2).value
    next_header = sheet.cell(row + 1, 2).value
    if not isinstance(header, str):
        return False
    return str(next_header).strip().lower() == "keterangan"


def parse_transactions(sheet) -> pd.DataFrame:
    records = []
    row = 1
    while row <= sheet.max_row - 1:
        if section_start(sheet, row):
            category = str(sheet.cell(row, 2).value).strip()
            row += 3

            while row <= sheet.max_row:
                if section_start(sheet, row):
                    break

                description = sheet.cell(row, 2).value
                date_value = sheet.cell(row, 3).value
                debit = to_number(sheet.cell(row, 5).value)

                if description and date_value and debit > 0:
                    records.append(
                        {
                            "Category": category,
                            "Date": pd.to_datetime(date_value),
                            "Description": str(description).strip(),
                            "Amount": debit,
                        }
                    )
                row += 1
            continue
        row += 1

    if not records:
        return pd.DataFrame(columns=["Category", "Date", "Description", "Amount"])

    table = pd.DataFrame(records)
    table.sort_values(by="Date", ascending=False, inplace=True)
    return table


def read_month_summary(month_sheet) -> dict[str, float]:
    return {
        "income": to_number(month_sheet["D2"].value),
        "planned_opex": to_number(month_sheet["D3"].value),
        "planned_savings": to_number(month_sheet["D4"].value),
        "planned_end_balance": to_number(month_sheet["D5"].value),
    }


def build_scenarios(income: float, base_opex: float) -> pd.DataFrame:
    scenario_rows = [
        {"Scenario": "Conservative Growth", "SchoolOpex": 1_500_000},
        {"Scenario": "Balanced Excellence", "SchoolOpex": 2_500_000},
        {"Scenario": "Stretch Premium", "SchoolOpex": 3_500_000},
    ]

    rows = []
    for row in scenario_rows:
        school_opex = float(row["SchoolOpex"])
        total_opex = base_opex + school_opex
        savings = income - total_opex
        savings_ratio = (savings / income) if income > 0 else 0
        edu_ratio = (school_opex / income) if income > 0 else 0

        if savings_ratio >= 0.25:
            health = "Very Safe"
        elif savings_ratio >= 0.15:
            health = "Manageable"
        else:
            health = "Tight"

        rows.append(
            {
                "Scenario": row["Scenario"],
                "School Opex": school_opex,
                "Total Opex": total_opex,
                "Savings": savings,
                "Savings Ratio": savings_ratio,
                "Education Ratio": edu_ratio,
                "Health": health,
            }
        )

    table = pd.DataFrame(rows)
    return table


def recommend_scenario(scenarios: pd.DataFrame) -> str:
    candidates = scenarios[
        (scenarios["Education Ratio"] >= 0.15)
        & (scenarios["Education Ratio"] <= 0.25)
        & (scenarios["Savings Ratio"] >= 0.20)
    ]
    if not candidates.empty:
        return str(candidates.iloc[0]["Scenario"])

    safe_candidates = scenarios[scenarios["Health"] != "Tight"]
    if not safe_candidates.empty:
        return str(safe_candidates.iloc[-1]["Scenario"])

    return str(scenarios.iloc[0]["Scenario"])


st.title("Preschool Budget and School Cost Dashboard")
st.caption(
    "Tracks monthly family spending from your workbook and tests TK-A plus TK-B cost scenarios with transport from Galaxy Bekasi."
)

with st.sidebar:
    st.header("Workbook Source")
    source_mode = st.radio(
        "Choose data source",
        ["Upload file", "Local file path", "Google Drive share link"],
        index=0,
    )

    workbook_bytes = None

    if source_mode == "Upload file":
        upload = st.file_uploader("Upload .xlsx", type=["xlsx"])
        if upload is not None:
            workbook_bytes = upload.read()

    elif source_mode == "Local file path":
        local_path = st.text_input("Absolute path to .xlsx", value="Pengeluaran_budget_template (1).xlsx")
        if local_path:
            try:
                with open(local_path, "rb") as handle:
                    workbook_bytes = handle.read()
            except FileNotFoundError:
                st.warning("Local file not found. Check path and try again.")

    else:
        drive_url = st.text_input("Google Drive shared file URL")
        if drive_url:
            try:
                workbook_bytes = download_drive_xlsx(drive_url)
            except Exception as err:
                st.error(f"Could not download from Google Drive: {err}")

if workbook_bytes is None:
    st.info("Load a workbook first to unlock the dashboard.")
    st.stop()

try:
    workbook = load_workbook(workbook_bytes)
except Exception as err:
    st.error(f"Workbook could not be read: {err}")
    st.stop()

month_options = get_month_sheets(workbook)
if not month_options:
    st.error("No monthly sheet found. Expected at least one sheet with Income in cell B2.")
    st.stop()

selected_month = st.selectbox("Select month sheet", options=month_options, index=0)
month_sheet = workbook[selected_month]
month_summary = read_month_summary(month_sheet)
budget_table = parse_budget_table(workbook)
transactions = parse_transactions(month_sheet)

k1, k2, k3, k4 = st.columns(4)
k1.markdown(
    f"<div class='kpi-card'><div class='kpi-label'>Income</div><div class='kpi-value'>{idr(month_summary['income'])}</div><div class='kpi-hint'>From D2</div></div>",
    unsafe_allow_html=True,
)
k2.markdown(
    f"<div class='kpi-card'><div class='kpi-label'>Planned Opex</div><div class='kpi-value'>{idr(month_summary['planned_opex'])}</div><div class='kpi-hint'>From D3</div></div>",
    unsafe_allow_html=True,
)
k3.markdown(
    f"<div class='kpi-card'><div class='kpi-label'>Planned Savings</div><div class='kpi-value'>{idr(month_summary['planned_savings'])}</div><div class='kpi-hint'>From D4</div></div>",
    unsafe_allow_html=True,
)
k4.markdown(
    f"<div class='kpi-card'><div class='kpi-label'>Planned Ending Balance</div><div class='kpi-value'>{idr(month_summary['planned_end_balance'])}</div><div class='kpi-hint'>From D5</div></div>",
    unsafe_allow_html=True,
)

st.markdown("### Category Budget vs Realization")
if budget_table.empty:
    st.warning("Budget Template table not found in workbook.")
else:
    chart_source = budget_table.melt(
        id_vars=["Category"],
        value_vars=["Budget", "Actual"],
        var_name="Metric",
        value_name="Amount",
    )
    fig = px.bar(
        chart_source,
        x="Category",
        y="Amount",
        color="Metric",
        barmode="group",
        color_discrete_sequence=["#0f766e", "#f59e0b"],
        title=f"{selected_month}: Budget vs Actual",
    )
    fig.update_layout(xaxis_title="Category", yaxis_title="Amount (IDR)")
    st.plotly_chart(fig, use_container_width=True)

    table_view = budget_table.copy()
    table_view["Budget"] = table_view["Budget"].map(idr)
    table_view["Actual"] = table_view["Actual"].map(idr)
    table_view["Remaining"] = table_view["Remaining"].map(idr)
    table_view["UsedPct"] = (budget_table["UsedPct"] * 100).map(lambda v: f"{v:.1f}%")
    st.dataframe(table_view, use_container_width=True, hide_index=True)

st.markdown("### Transaction Feed")
if transactions.empty:
    st.info("No debit transactions were detected for this month sheet.")
else:
    all_categories = sorted(transactions["Category"].unique().tolist())
    selected_categories = st.multiselect(
        "Filter categories",
        options=all_categories,
        default=all_categories,
    )
    filtered = transactions[transactions["Category"].isin(selected_categories)].copy()
    spend_by_category = filtered.groupby("Category", as_index=False)["Amount"].sum().sort_values("Amount", ascending=False)

    t1, t2 = st.columns([1.1, 1.4])
    with t1:
        pie = px.pie(
            spend_by_category,
            names="Category",
            values="Amount",
            hole=0.52,
            color_discrete_sequence=px.colors.sequential.Tealgrn,
            title="Spend mix by category",
        )
        st.plotly_chart(pie, use_container_width=True)

    with t2:
        table = filtered.copy()
        table["Date"] = table["Date"].dt.strftime("%Y-%m-%d")
        table["Amount"] = table["Amount"].map(idr)
        st.dataframe(table[["Date", "Category", "Description", "Amount"]], use_container_width=True, hide_index=True)

st.markdown("### Preschool Scenario Planner")
st.markdown("Use this to test affordability for TK-A and TK-B while protecting family savings.")

s1, s2 = st.columns(2)
with s1:
    income_input = st.number_input(
        "Monthly take-home income (IDR)",
        min_value=0,
        step=100_000,
        value=int(month_summary["income"] or 0),
    )
with s2:
    base_opex_input = st.number_input(
        "Current monthly household opex without preschool (IDR)",
        min_value=0,
        step=100_000,
        value=int(month_summary["planned_opex"] or 0),
    )

scenario_table = build_scenarios(float(income_input), float(base_opex_input))
recommended = recommend_scenario(scenario_table)

scenario_view = scenario_table.copy()
scenario_view["School Opex"] = scenario_view["School Opex"].map(idr)
scenario_view["Total Opex"] = scenario_view["Total Opex"].map(idr)
scenario_view["Savings"] = scenario_view["Savings"].map(idr)
scenario_view["Savings Ratio"] = (scenario_table["Savings Ratio"] * 100).map(lambda v: f"{v:.1f}%")
scenario_view["Education Ratio"] = (scenario_table["Education Ratio"] * 100).map(lambda v: f"{v:.1f}%")

st.dataframe(scenario_view, use_container_width=True, hide_index=True)
st.success(f"Recommended starting scenario: {recommended}")

st.markdown(
    """
<div class='small-muted'>
Formula reminder for 2-year preschool cost:<br>
Two-year total = (Monthly Opex x 24) + (Annual Fees x 2) + One-time fees.
</div>
""",
    unsafe_allow_html=True,
)
