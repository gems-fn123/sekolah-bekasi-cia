from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


def to_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace("Rp", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def get_month_sheets(workbook) -> list[str]:
    months: list[str] = []
    for sheet in workbook.worksheets:
        title = sheet.title.strip().lower()
        if title == "budget template":
            continue
        if sheet.max_row <= 1:
            continue
        if str(sheet["B2"].value).strip().lower() == "income":
            months.append(sheet.title)
    return months


def parse_budget_template(workbook) -> pd.DataFrame:
    if "Budget Template" not in workbook.sheetnames:
        return pd.DataFrame()

    sheet = workbook["Budget Template"]
    header_row = None
    for row in range(1, min(sheet.max_row, 80) + 1):
        if str(sheet.cell(row, 1).value).strip().lower() == "kategori":
            header_row = row
            break

    if header_row is None:
        return pd.DataFrame()

    rows = []
    for row in range(header_row + 1, sheet.max_row + 1):
        category = sheet.cell(row, 1).value
        if category is None:
            continue
        if str(category).strip().upper() == "TOTAL":
            break
        rows.append(
            {
                "category": str(category).strip(),
                "budget": to_number(sheet.cell(row, 2).value),
                "actual": to_number(sheet.cell(row, 3).value),
                "remaining": to_number(sheet.cell(row, 4).value),
                "used_pct": to_number(sheet.cell(row, 5).value),
                "status": str(sheet.cell(row, 6).value or "").strip(),
                "type": str(sheet.cell(row, 7).value or "").strip(),
                "notes": str(sheet.cell(row, 8).value or "").strip(),
            }
        )

    return pd.DataFrame(rows)


def section_start(sheet, row: int) -> bool:
    if row + 1 > sheet.max_row:
        return False
    header = sheet.cell(row, 2).value
    next_header = sheet.cell(row + 1, 2).value
    if not isinstance(header, str):
        return False
    return str(next_header).strip().lower() == "keterangan"


def parse_transactions(month_name: str, sheet) -> pd.DataFrame:
    rows = []
    row = 1
    while row <= sheet.max_row - 1:
        if section_start(sheet, row):
            category = str(sheet.cell(row, 2).value).strip()
            row += 3
            while row <= sheet.max_row:
                if section_start(sheet, row):
                    break
                desc = sheet.cell(row, 2).value
                date_value = sheet.cell(row, 3).value
                debit = to_number(sheet.cell(row, 5).value)

                if desc and date_value and debit > 0:
                    rows.append(
                        {
                            "month_sheet": month_name,
                            "category": category,
                            "date": pd.to_datetime(date_value).date().isoformat(),
                            "description": str(desc).strip(),
                            "amount": debit,
                        }
                    )
                row += 1
            continue
        row += 1

    return pd.DataFrame(rows)


def parse_month_summary(month_name: str, sheet) -> dict[str, float | str]:
    return {
        "month_sheet": month_name,
        "income": to_number(sheet["D2"].value),
        "planned_opex": to_number(sheet["D3"].value),
        "planned_savings": to_number(sheet["D4"].value),
        "planned_end_balance": to_number(sheet["D5"].value),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Export workbook tables for Google Sheets and Looker Studio")
    parser.add_argument("--input", required=True, help="Path to budget workbook (.xlsx)")
    parser.add_argument("--output", default="dashboard/exports", help="Output directory for CSV exports")
    parser.add_argument(
        "--schools",
        default="outputs/tables/preschool_rankings_comprehensive.csv",
        help="Optional school ranking/candidate CSV to copy into exports",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.load_workbook(input_path, data_only=True)

    months = get_month_sheets(workbook)
    if not months:
        raise ValueError("No monthly sheet found. Expected Income in B2 on at least one sheet.")

    budget = parse_budget_template(workbook)
    if not budget.empty:
        budget.to_csv(output_dir / "budget_categories.csv", index=False)

    summary_rows = []
    tx_frames = []
    for month in months:
        sheet = workbook[month]
        summary_rows.append(parse_month_summary(month, sheet))
        tx_frame = parse_transactions(month, sheet)
        if not tx_frame.empty:
            tx_frames.append(tx_frame)

    summary = pd.DataFrame(summary_rows)
    summary.to_csv(output_dir / "monthly_summary.csv", index=False)

    if tx_frames:
        transactions = pd.concat(tx_frames, ignore_index=True)
    else:
        transactions = pd.DataFrame(columns=["month_sheet", "category", "date", "description", "amount"])
    transactions.to_csv(output_dir / "transactions.csv", index=False)

    school_path = Path(args.schools)
    if school_path.exists():
        schools = pd.read_csv(school_path)
        schools.to_csv(output_dir / "school_candidates.csv", index=False)

    print(f"Export complete: {output_dir.resolve()}")


if __name__ == "__main__":
    main()
